"""
NCRB 'Crime in India 2023' XLSX -> normalized long-format JSONL.

Source: OpenCity India (https://data.opencity.in/dataset/crime-in-india-2023)
Original publisher: National Crime Records Bureau (NCRB), Government of India.

Output schema (one record per row):
    {
      "state":         str,          # mapped from city -> state (best-effort)
      "city":          str,          # 'Bengaluru', 'Delhi City', ...
      "district":      str | None,   # NCRB metro tables = city-level; no sub-district
      "crime_type":    str,          # leaf metric label, lower-snake
      "metric_path":   str,          # full pipe-joined header path for traceability
      "year":          int,          # 2023
      "count":         float | int,  # numeric value
      "unit":          str | None,   # 'cases' | 'kg' | 'litre' | 'numbers' | 'crores' | None
      "source_file":   str,          # original filename
      "ncrb_table_id": str | None,   # if inferable from title, else None
      "is_total_row":  bool          # true for the '53 metropolitan cities total' row
    }

[UNCERTAIN] markers appear where the source format itself is ambiguous.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "raw"
OUT_ALL = ROOT / "all_states.jsonl"
OUT_KA = ROOT / "karnataka_aggregates.jsonl"

# 53-Metropolitan-City -> State mapping (NCRB 2023 metro list).
CITY_TO_STATE: dict[str, str] = {
    "Agra": "Uttar Pradesh",
    "Ahmedabad": "Gujarat",
    "Amritsar": "Punjab",
    "Asansol": "West Bengal",
    "Aurangabad": "Maharashtra",
    "Bengaluru": "Karnataka",
    "Bhopal": "Madhya Pradesh",
    "Chandigarh City": "Chandigarh",
    "Chennai": "Tamil Nadu",
    "Coimbatore": "Tamil Nadu",
    "Delhi City": "Delhi",
    "Dhanbad": "Jharkhand",
    "Durg-Bhilainagar": "Chhattisgarh",
    "Faridabad": "Haryana",
    "Ghaziabad": "Uttar Pradesh",
    "Gwalior": "Madhya Pradesh",
    "Hyderabad": "Telangana",
    "Indore": "Madhya Pradesh",
    "Jabalpur": "Madhya Pradesh",
    "Jaipur": "Rajasthan",
    "Jamshedpur": "Jharkhand",
    "Jodhpur": "Rajasthan",
    "Kannur": "Kerala",
    "Kanpur": "Uttar Pradesh",
    "Kochi": "Kerala",
    "Kolkata": "West Bengal",
    "Kollam": "Kerala",
    "Kota": "Rajasthan",
    "Kozhikode": "Kerala",
    "Lucknow": "Uttar Pradesh",
    "Ludhiana": "Punjab",
    "Madurai": "Tamil Nadu",
    "Malappuram": "Kerala",
    "Meerut": "Uttar Pradesh",
    "Mumbai": "Maharashtra",
    "Nagpur": "Maharashtra",
    "Nasik": "Maharashtra",
    "Patna": "Bihar",
    "Prayagraj": "Uttar Pradesh",
    "Pune": "Maharashtra",
    "Raipur": "Chhattisgarh",
    "Rajkot": "Gujarat",
    "Ranchi": "Jharkhand",
    "Srinagar": "Jammu and Kashmir",
    "Surat": "Gujarat",
    "Thiruvananthapuram": "Kerala",
    "Thrissur": "Kerala",
    "Tiruchirapalli": "Tamil Nadu",
    "Vadodara": "Gujarat",
    "Varanasi": "Uttar Pradesh",
    "Vasai Virar": "Maharashtra",
    "Vijayawada": "Andhra Pradesh",
    "Vishakhapatnam": "Andhra Pradesh",
}


def to_snake(s: str) -> str:
    s = re.sub(r"[^0-9a-zA-Z]+", "_", s.strip().lower())
    return re.sub(r"_+", "_", s).strip("_")


def num(v):
    """Coerce numeric cells; return None if not numeric."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def is_total_row(name: str) -> bool:
    return bool(name) and "total" in name.lower() and "cities" in name.lower()


def flatten_headers(ws, header_rows: list[int]) -> list[str]:
    """
    Forward-fill merged-cell headers across `header_rows` and pipe-join into
    one leaf label per data column. Returns a list with one entry per column.
    """
    max_col = ws.max_column
    grid: list[list[str | None]] = []
    for r in header_rows:
        row: list[str | None] = []
        last = None
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                v = last
            else:
                v = str(v).strip()
                last = v
            row.append(v)
        grid.append(row)

    labels: list[str] = []
    for c in range(max_col):
        parts: list[str] = []
        seen: set[str] = set()
        for r in range(len(header_rows)):
            cell = grid[r][c]
            if cell and cell not in seen:
                parts.append(cell)
                seen.add(cell)
        labels.append(" | ".join(parts) if parts else f"col_{c+1}")
    return labels


def emit_rows(
    xlsx_path: Path,
    header_rows: list[int],
    data_start_row: int,
    city_col: int = 2,
    table_id: str | None = None,
    default_unit: str | None = "cases",
):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["report"]
    labels = flatten_headers(ws, header_rows)

    for r in range(data_start_row, ws.max_row + 1):
        city = ws.cell(row=r, column=city_col).value
        if not city or not isinstance(city, str):
            continue
        city = city.strip()
        total_flag = is_total_row(city)
        state = (
            "ALL_INDIA_METRO_TOTAL"
            if total_flag
            else CITY_TO_STATE.get(city, "[UNCERTAIN]")
        )

        for c in range(1, ws.max_column + 1):
            # Skip the SL-No column and the city-name column.
            if c <= city_col:
                continue
            val = ws.cell(row=r, column=c).value
            n = num(val)
            if n is None:
                continue
            label = labels[c - 1]
            # Unit detection from label
            unit = default_unit
            ll = label.lower()
            if "crore" in ll:
                unit = "crores_inr"
            elif "in kg" in ll or "(in kg)" in ll:
                unit = "kg"
            elif "in litre" in ll or "(in litre)" in ll:
                unit = "litre"
            elif "in numbers" in ll or "(in numbers)" in ll:
                unit = "numbers"

            yield {
                "state": state,
                "city": city,
                "district": None,
                "crime_type": to_snake(label.split(" | ")[-1])[:120],
                "metric_path": label,
                "year": 2023,
                "count": n,
                "unit": unit,
                "source_file": xlsx_path.name,
                "ncrb_table_id": table_id,
                "is_total_row": total_flag,
            }


# Per-file extraction plans (header-row positions derived from probing).
PLANS = [
    {
        "file": "ipc_156_3_2023.xlsx",
        "header_rows": [2, 3],
        "data_start": 4,
        "table_id": "NCRB-2023-CITY-IPC-156_3",
        "default_unit": "cases",
    },
    {
        "file": "persons_disposal_ipc_2023.xlsx",
        "header_rows": [2, 3, 4, 5],
        "data_start": 7,
        "table_id": "NCRB-2023-CITY-PERSONS-DISPOSAL-IPC",
        "default_unit": "persons",
    },
    {
        "file": "persons_disposal_women_2023.xlsx",
        "header_rows": [2, 3, 4, 5],
        "data_start": 7,
        "table_id": "NCRB-2023-CITY-PERSONS-DISPOSAL-WOMEN",
        "default_unit": "persons",
    },
    {
        "file": "persons_disposal_children_2023.xlsx",
        "header_rows": [2, 3, 4, 5],
        "data_start": 7,
        "table_id": "NCRB-2023-CITY-PERSONS-DISPOSAL-CHILDREN",
        "default_unit": "persons",
    },
    {
        "file": "property_stolen_recovered_2023.xlsx",
        "header_rows": [2, 3, 4],
        "data_start": 6,
        "table_id": "NCRB-2023-CITY-PROPERTY-STOLEN-RECOVERED",
        "default_unit": "cases",
    },
    {
        "file": "ndps_seizures_2023.xlsx",
        "header_rows": [2, 3, 4],
        "data_start": 6,
        "table_id": "NCRB-2023-CITY-NDPS-SEIZURES",
        "default_unit": None,  # heterogeneous units (kg/litre/numbers) per column
    },
]


def main() -> None:
    all_records = []
    summary = []
    for plan in PLANS:
        path = RAW / plan["file"]
        if not path.exists():
            print(f"  SKIP (missing): {path.name}")
            continue
        before = len(all_records)
        for rec in emit_rows(
            path,
            plan["header_rows"],
            plan["data_start"],
            table_id=plan["table_id"],
            default_unit=plan["default_unit"],
        ):
            all_records.append(rec)
        added = len(all_records) - before
        summary.append((plan["file"], added))
        print(f"  {plan['file']:<42} +{added:>5} records")

    print(f"\nTotal records: {len(all_records)}")
    print(f"Karnataka records: {sum(1 for r in all_records if r['state'] == 'Karnataka')}")

    with OUT_ALL.open("w", encoding="utf-8") as f:
        for r in all_records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    with OUT_KA.open("w", encoding="utf-8") as f:
        for r in all_records:
            if r["state"] == "Karnataka":
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"\nWrote {OUT_ALL.name} and {OUT_KA.name}")


if __name__ == "__main__":
    main()
